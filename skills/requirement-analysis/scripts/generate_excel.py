#!/usr/bin/env python3
"""
Excel 测试用例生成器
读取 JSON 格式的测试用例数据，生成格式化的 .xlsx 文件。
"""

import json
import os
import argparse
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ============================================================
# 样式定义
# ============================================================

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=10)
TITLE_FONT = Font(name="微软雅黑", size=14, bold=True)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# 用例分级颜色（匹配飞书模板）
PRIORITY_FILLS = {
    "P0": PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),  # 红色 — 冒烟用例
    "P1": PatternFill(start_color="FFE6A0", end_color="FFE6A0", fill_type="solid"),  # 橙色 — 其他功能测试用例
    "P2": PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),  # 绿色 — 页面样式校验用例
}

PRIORITY_FONTS = {
    "P0": Font(name="微软雅黑", size=10, color="9C0006", bold=True),
    "P1": Font(name="微软雅黑", size=10, color="9C6500", bold=True),
    "P2": Font(name="微软雅黑", size=10, color="006100"),
}

WRAP_ALIGNMENT = Alignment(
    wrap_text=True,
    vertical="top",
    horizontal="left",
)
CENTER_ALIGNMENT = Alignment(
    wrap_text=True,
    vertical="top",
    horizontal="center",
)


# ============================================================
# 列定义
# ============================================================

# 默认列配置 — 与飞书测试用例管理模板一致
DEFAULT_COLUMNS = [
    {"field": "name", "header": "用例名称", "width": 40},
    {"field": "directory", "header": "所属目录", "width": 36},
    {"field": "steps", "header": "执行步骤", "width": 44},
    {"field": "type", "header": "用例类型", "width": 22},
    {"field": "expected", "header": "预期结果", "width": 40},
    {"field": "precondition", "header": "前置条件", "width": 26},
    {"field": "requirement_id", "header": "关联需求", "width": 16},
    {"field": "priority", "header": "用例分级", "width": 26},
    {"field": "owner", "header": "负责人", "width": 12},
]

# 哪些列为居中对齐
CENTER_COLUMNS = {"type", "priority", "owner"}


def load_columns_from_config(config: dict) -> list[dict]:
    """从配置文件中加载列定义"""
    template = config.get("testcase_template", {})
    col_names = template.get("columns", [])
    if not col_names:
        return DEFAULT_COLUMNS

    # 从列名列表构建列定义
    columns = []
    for name in col_names:
        col = {
            "field": _name_to_field(name),
            "header": name,
            "width": 20,  # 默认宽度
        }
        # 调整常见列的宽度
        if "步骤" in name:
            col["width"] = 44
        elif "预期" in name or "结果" in name:
            col["width"] = 40
        elif "名称" in name or "标题" in name:
            col["width"] = 40
        elif "目录" in name:
            col["width"] = 36
        elif "分级" in name or "级别" in name:
            col["width"] = 26
        elif "类型" in name:
            col["width"] = 22
        elif "需求" in name:
            col["width"] = 16
        elif "负责人" in name:
            col["width"] = 12
        columns.append(col)
    return columns


def _name_to_field(name: str) -> str:
    """列名 → JSON 字段名映射（飞书测试用例管理模板）"""
    mapping = {
        "用例名称": "name",
        "用例标题": "name",
        "测试标题": "name",
        "所属目录": "directory",
        "目录": "directory",
        "模块": "directory",
        "执行步骤": "steps",
        "测试步骤": "steps",
        "操作步骤": "steps",
        "步骤": "steps",
        "用例类型": "type",
        "测试类型": "type",
        "类型": "type",
        "预期结果": "expected",
        "期望结果": "expected",
        "预期": "expected",
        "前置条件": "precondition",
        "前置": "precondition",
        "关联需求": "requirement_id",
        "需求ID": "requirement_id",
        "用例分级": "priority",
        "优先级": "priority",
        "负责人": "owner",
        "处理人": "owner",
    }
    return mapping.get(name, name.lower())


# ============================================================
# 数据加载
# ============================================================

def load_testcases(input_path: str) -> list[dict]:
    """加载测试用例 JSON 数据"""
    with open(input_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, list):
        raise ValueError("测试用例数据必须是 JSON 数组")
    return data


# ============================================================
# Excel 生成
# ============================================================

def generate_excel(
    testcases: list[dict],
    columns: list[dict],
    output_path: str,
    metadata: dict = None,
) -> str:
    """
    生成格式化的 Excel 文件。

    Args:
        testcases: 测试用例数据列表
        columns: 列定义 [{"field": ..., "header": ..., "width": ...}]
        output_path: 输出文件路径
        metadata: 可选的元数据（来源、生成时间等），写入 Sheet 顶部信息区

    Returns:
        输出文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "测试用例"

    current_row = 1

    # --- 元数据信息行 ---
    if metadata:
        info_lines = []
        if metadata.get("title"):
            info_lines.append(metadata["title"])
        info_lines.append(f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        if metadata.get("source"):
            info_lines.append(f"需求来源：{metadata['source']}")
        total = len(testcases)
        p0 = sum(1 for t in testcases if "P0" in str(t.get("priority", "")))
        p1 = sum(1 for t in testcases if "P1" in str(t.get("priority", "")))
        p2 = sum(1 for t in testcases if "P2" in str(t.get("priority", "")))
        info_lines.append(f"用例统计：共 {total} 条 | P0={p0} P1={p1} P2={p2}")

        for line in info_lines:
            cell = ws.cell(row=current_row, column=1, value=line)
            cell.font = Font(name="微软雅黑", size=10, color="666666")
            ws.merge_cells(
                start_row=current_row,
                start_column=1,
                end_row=current_row,
                end_column=len(columns),
            )
            current_row += 1
        current_row += 1  # 空一行

    # --- 表头行 ---
    header_row = current_row
    for col_idx, col_def in enumerate(columns, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=col_def["header"])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER_ALIGNMENT
        cell.border = THIN_BORDER
    current_row += 1

    # --- 数据行 ---
    for tc in testcases:
        for col_idx, col_def in enumerate(columns, 1):
            value = tc.get(col_def["field"], "")
            cell = ws.cell(row=current_row, column=col_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER

            if col_def["field"] in CENTER_COLUMNS:
                cell.alignment = CENTER_ALIGNMENT
            else:
                cell.alignment = WRAP_ALIGNMENT

            # 优先级列特殊染色
            if col_def["field"] == "priority":
                for prefix, fill in PRIORITY_FILLS.items():
                    if str(value).startswith(prefix):
                        cell.fill = fill
                        cell.font = PRIORITY_FONTS.get(prefix, BODY_FONT)
                        break

        current_row += 1

    # --- 列宽 ---
    for col_idx, col_def in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = col_def.get("width", 20)

    # --- 冻结首行 ---
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    # --- 自动筛选 ---
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(len(columns))}{current_row - 1}"

    # --- 行高 ---
    ws.row_dimensions[header_row].height = 28
    for r in range(header_row + 1, current_row):
        ws.row_dimensions[r].height = 60

    # --- 保存 ---
    output_dir = os.path.dirname(output_path)
    if output_dir:
        os.makedirs(output_dir, exist_ok=True)
    wb.save(output_path)
    return output_path


def print_summary(testcases: list[dict]):
    """打印测试用例摘要"""
    total = len(testcases)
    p0 = sum(1 for t in testcases if "P0" in str(t.get("priority", "")))
    p1 = sum(1 for t in testcases if "P1" in str(t.get("priority", "")))
    p2 = sum(1 for t in testcases if "P2" in str(t.get("priority", "")))

    # 按用例类型统计
    types: dict = {}
    for t in testcases:
        ty = t.get("type", "未分类")
        types[ty] = types.get(ty, 0) + 1

    # 按目录统计
    directories: dict = {}
    for t in testcases:
        d = t.get("directory", "未分类")
        directories[d] = directories.get(d, 0) + 1

    print(f"📊 用例统计：")
    print(f"   总计：{total} 条")
    print(f"   P0（冒烟用例）：{p0} 条")
    print(f"   P1（其他功能测试用例）：{p1} 条")
    print(f"   P2（页面样式校验用例）：{p2} 条")
    print(f"📂 用例类型分布：")
    for ty, count in sorted(types.items(), key=lambda x: -x[1]):
        print(f"   {ty}：{count} 条")
    print(f"📁 目录分布：")
    for d, count in sorted(directories.items(), key=lambda x: -x[1]):
        print(f"   {d}：{count} 条")


# ============================================================
# CLI
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Excel 测试用例生成器")
    parser.add_argument("--input", required=True, help="测试用例 JSON 文件路径")
    parser.add_argument("--output", default=None, help="输出 Excel 文件路径")
    parser.add_argument("--config", default=None, help="配置文件路径（用于读取列定义）")
    parser.add_argument("--columns", default=None, help="自定义列定义 JSON 文件路径（优先级高于 config）")

    args = parser.parse_args()

    # 加载测试用例数据
    testcases = load_testcases(args.input)

    # 加载列定义
    if args.columns:
        with open(args.columns, "r", encoding="utf-8") as f:
            columns = json.load(f)
    elif args.config:
        config = load_config(args.config)
        columns = load_columns_from_config(config)
    else:
        columns = DEFAULT_COLUMNS

    # 生成输出路径
    if args.output is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        args.output = os.path.join(
            os.path.dirname(args.input) or ".",
            f"测试用例_{timestamp}.xlsx",
        )

    # 生成元数据
    metadata = {
        "title": "测试用例",
        "source": args.input,
    }

    # 生成 Excel
    output_path = generate_excel(testcases, columns, args.output, metadata)

    print(f"✅ Excel 测试用例生成完成！")
    print(f"📄 文件位置：{os.path.abspath(output_path)}")
    print_summary(testcases)


def load_config(config_path: str) -> dict:
    """加载配置文件"""
    with open(config_path, "r", encoding="utf-8") as f:
        return json.load(f)


if __name__ == "__main__":
    main()
